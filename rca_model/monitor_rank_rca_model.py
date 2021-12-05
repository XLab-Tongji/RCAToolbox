import numpy as np
from openpyxl import load_workbook, Workbook
import os
from base.base_rca_model import BaseRCAModel
from utils.pearson import calc_pearson
from utils.build_graph import build_graph_pc


class MonitorRankRCAModel(BaseRCAModel):
    """
   MonitorRank根因分析
    """

    @staticmethod
    def normalize(p):
        """
        Normalize the matrix in each row
        """
        p = p.copy()
        for i in range(p.shape[0]):
            row_sum = np.sum(p[i])
            if row_sum > 0:
                p[i] /= row_sum
        return p

    def rela_to_rank(self, rela, call_graph, frontend, rho=0.1):
        n = len(call_graph)

        s = [abs(_) for _ in rela[frontend - 1]]

        p = np.zeros([n, n])
        for i in range(n):
            for j in range(n):
                # forward edge
                if call_graph[i][j] != 0:
                    p[i, j] = abs(s[j])
                # backward edge
                elif call_graph[j][i] != 0:
                    p[i, j] = rho * abs(s[i])
        # Add self edges
        for i in range(n):
            if i != frontend - 1:
                p[i][i] = max(0, s[i] - max(p[i]))
        teleportation_prob = (np.array(s) / np.sum(s)).tolist()
        p = self.normalize(p)
        return p, teleportation_prob

    @staticmethod
    def get_metric_data(experiment):
        """
         获得某次实验的metric数据
         :param experiment:单次实验数据
         :return:存储微服务名的列表和对应下标的微服务运行数据（对过程中数据全为0的微服务进行了删除)
          """
        data = [experiment['metric'][i].sample['value']for i in range(len(experiment['metric']))]
        data = np.array(data)
        header = [experiment['metric'][i].name for i in range(len(experiment['metric']))]
        idx = np.argwhere(np.all(data[..., :] == 0, axis=1))
        data = np.delete(data, idx, axis=0)
        header = np.delete(header, idx, axis=0)
        return header, data

    def build(self, train_data, config):
        """
        使用monitor_rank框架构建rca模型
        :param train_data: 训练数据，与base/base_data_loader.py中读入的train_data格式一致.
        :param config: 模型参数.
        :return 如果是每组实验一个模型，返回一个dict，key为experiment_id；如果是训练集整个是一个模型，返回该模型.
        """
        model = dict()
        for experiment_id in train_data.keys():
            header, data = self.get_metric_data(train_data[experiment_id])
            rela = calc_pearson(data, method="numpy", zero_diag=False)  # 皮尔森相关系数计算,返回一个实验对应一个相关系数矩阵
            base_dir = str(os.path.dirname(os.path.dirname(__file__)))
            if os.path.exists(base_dir+'/saved/model/monitor_rank_runner/'+experiment_id + '.xlsx'):
             wb = load_workbook(base_dir+'/saved/model/monitor_rank_runner/'+experiment_id+'.xlsx')
             sheets = wb.worksheets
             # 获取第一张sheet
             sheet1 = sheets[0]
             call_graph = []
             for row in sheet1:
                 row_data = []
                 for cell in row:
                     if cell.value is not None:
                         row_data.append(cell.value)
                 call_graph.append(row_data)
            else:
                call_graph = build_graph_pc(data, config['alpha'])
                call_graph = np.array(call_graph)
                # 创建一个Workbook对象
                wb = Workbook()
                # 获取当前活跃的sheet，默认是第一个sheet
                ws = wb.active
                for i in range(0, len(call_graph)):
                    for j in range(0, len(call_graph[i])):
                        ws.cell(i+1, j+1).value = call_graph[i][j]
                wb.save(base_dir+'/data/demo/metric/dep_graph/'+experiment_id + '.xlsx')

            p, teleportation_prob = self.rela_to_rank(rela, call_graph, config['frontend'])
            out = dict()
            out['header'] = header
            out['pc_graph'] = p
            out['teleportation_prob'] = teleportation_prob
            model[experiment_id] = out
        return model
